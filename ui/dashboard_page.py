"""
ui/dashboard_page.py  —  Phase Chart (Dialog)

Dashboard tổng quan cho gym management:
  • StatCard           — card hiển thị 1 chỉ số
  • TodayDetailTable   — bảng danh sách check-in hôm nay theo member
  • TopMembersTable    — bảng top 5 đi tập nhiều nhất tháng
  • MemberHistoryPanel — combobox + bảng lịch sử 1 hội viên
  • ChartDialog        — dialog popup hiển thị Pie + Bar chart
  • DashboardPage      — trang tổng hợp
"""

from __future__ import annotations

import matplotlib
matplotlib.use("Qt5Agg")

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from datetime import date
from ui.chart_widget import PieChartWidget, BarChartWidget
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt, QTimer, Slot
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from controllers.dashboard_controller import DashboardController


# ══════════════════════════════════════════════════════════════════════════ #
#  StatCard                                                                   #
# ══════════════════════════════════════════════════════════════════════════ #

class StatCard(QFrame):
    """Hiển thị 1 chỉ số với icon + label + giá trị lớn."""

    def __init__(
        self,
        icon: str,
        label: str,
        value: str = "—",
        accent: str = "#3b82f6",
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setObjectName("statCard")
        self._accent = accent
        self._setup_ui(icon, label, value)

    def _setup_ui(self, icon: str, label: str, value: str) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 8, 10, 8)
        layout.setSpacing(2)

        top = QHBoxLayout()
        top.setSpacing(8)

        icon_lbl = QLabel(icon)
        icon_lbl.setObjectName("statCardIcon")
        top.addWidget(icon_lbl)

        lbl = QLabel(label)
        lbl.setObjectName("statCardLabel")
        lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        top.addWidget(lbl)
        top.addStretch()

        layout.addLayout(top)

        self._value_lbl = QLabel(value)
        self._value_lbl.setObjectName("statCardValue")
        self._value_lbl.setStyleSheet(f"color: {self._accent};")
        layout.addWidget(self._value_lbl)

    def set_value(self, value: str) -> None:
        self._value_lbl.setText(value)


# ══════════════════════════════════════════════════════════════════════════ #
#  Helpers                                                                    #
# ══════════════════════════════════════════════════════════════════════════ #

def _make_table(headers: list[str]) -> QTableWidget:
    tbl = QTableWidget(0, len(headers))
    tbl.setHorizontalHeaderLabels(headers)
    tbl.setEditTriggers(QTableWidget.NoEditTriggers)
    tbl.setSelectionBehavior(QTableWidget.SelectRows)
    tbl.setAlternatingRowColors(True)
    tbl.verticalHeader().setVisible(False)
    tbl.horizontalHeader().setHighlightSections(False)
    tbl.setShowGrid(False)
    tbl.setSortingEnabled(True)
    return tbl


def _cell(
    tbl: QTableWidget,
    row: int,
    col: int,
    text: str,
    align: Qt.AlignmentFlag = Qt.AlignLeft | Qt.AlignVCenter,
    color: QColor | None = None,
) -> QTableWidgetItem:
    item = QTableWidgetItem(str(text))
    item.setTextAlignment(align)
    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
    if color:
        item.setForeground(color)
    tbl.setItem(row, col, item)
    return item


def _section_title(text: str) -> QLabel:
    lbl = QLabel(text)
    lbl.setObjectName("sectionTitle")
    return lbl


# ══════════════════════════════════════════════════════════════════════════ #
#  TodayDetailTable                                                           #
# ══════════════════════════════════════════════════════════════════════════ #

class TodayDetailTable(QWidget):
    HEADERS = ["#", "Tên Hội Viên", "Số Lần Check-in Hôm Nay"]

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)
        layout.addWidget(_section_title("📋  Chi Tiết Check-in Hôm Nay"))
        self._table = _make_table(self.HEADERS)
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        layout.addWidget(self._table)

    def load(self, detail: list[dict]) -> None:
        self._table.setSortingEnabled(False)
        self._table.setRowCount(0)
        for i, row in enumerate(detail):
            r = self._table.rowCount()
            self._table.insertRow(r)
            _cell(self._table, r, 0, str(i + 1), Qt.AlignCenter)
            _cell(self._table, r, 1, row.get("member_name", ""))
            _cell(self._table, r, 2, str(row.get("count", 0)), Qt.AlignCenter)
        self._table.setSortingEnabled(False)
        self._table.sortItems(0, Qt.AscendingOrder)


# ══════════════════════════════════════════════════════════════════════════ #
#  TopMembersTable                                                            #
# ══════════════════════════════════════════════════════════════════════════ #

class TopMembersTable(QWidget):
    HEADERS = ["Hạng", "Tên Hội Viên", "Số Ngày Đi Tập (Tháng)"]
    MEDALS  = ["🥇", "🥈", "🥉", "4", "5"]

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)
        layout.addWidget(_section_title("🏆  Top 5 Hội Viên Tháng Này"))
        self._table = _make_table(self.HEADERS)
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        layout.addWidget(self._table)

    def load(self, rows: list[dict]) -> None:
        self._table.setSortingEnabled(False)
        self._table.setRowCount(0)

        rows = sorted(rows, key=lambda x: x.get("days_count", 0), reverse=True)

        medal_colors = [
            QColor("#f59e0b"),   # gold
            QColor("#94a3b8"),   # silver
            QColor("#cd7f32"),   # bronze
        ]
        for i, row in enumerate(rows):
            r = self._table.rowCount()
            self._table.insertRow(r)
            medal = self.MEDALS[i] if i < len(self.MEDALS) else str(i + 1)
            color = medal_colors[i] if i < 3 else None
            _cell(self._table, r, 0, medal, Qt.AlignCenter, color)
            _cell(self._table, r, 1, row.get("member_name", ""), color=color)
            _cell(self._table, r, 2, str(row.get("days_count", 0)), Qt.AlignCenter, color)

        self._table.setSortingEnabled(False)


# ══════════════════════════════════════════════════════════════════════════ #
#  MemberHistoryPanel                                                         #
# ══════════════════════════════════════════════════════════════════════════ #

class MemberHistoryPanel(QWidget):
    HEADERS = ["#", "Thời Gian Check-in", "Trạng Thái"]

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._members: list[dict] = []
        self._setup_ui()

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        layout.addWidget(_section_title("📅  Lịch Sử Check-in Theo Hội Viên"))

        self._combo = QComboBox()
        self._combo.setPlaceholderText("— Chọn hội viên để xem lịch sử —")
        self._combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self._combo.currentIndexChanged.connect(self._on_member_changed)
        layout.addWidget(self._combo)

        self._count_lbl = QLabel("")
        self._count_lbl.setObjectName("statusLabel")
        layout.addWidget(self._count_lbl)

        self._table = _make_table(self.HEADERS)
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)

        # Mở rộng cột thời gian check-in tối đa
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)

        layout.addWidget(self._table)

    def populate_members(self, members: list[dict]) -> None:
        self._members = members
        self._combo.blockSignals(True)
        self._combo.clear()
        for m in members:
            display = f"{m['name']}  ({m.get('phone', '')})"
            self._combo.addItem(display, userData=m["id"])
        self._combo.blockSignals(False)

    def load_history(self, rows: list[dict]) -> None:
        self._table.setSortingEnabled(False)
        self._table.setRowCount(0)
        for i, row in enumerate(rows):
            r = self._table.rowCount()
            self._table.insertRow(r)
            _cell(self._table, r, 0, str(i + 1), Qt.AlignCenter)

            # Căn trái thời gian check-in để dễ nhìn hơn khi cột được kéo rộng
            _cell(self._table, r, 1, row.get("checkin_time", ""), Qt.AlignLeft | Qt.AlignVCenter)

            status = row.get("status", "")
            if status == "valid":
                label, color = "✅  Hợp Lệ",      QColor("#22c55e")
            elif status == "expired":
                label, color = "⚠️  Hết Hạn",     QColor("#f59e0b")
            else:
                label, color = "❌  Chưa Đăng Ký", QColor("#ef4444")
            _cell(self._table, r, 2, label, Qt.AlignCenter, color)

        count = len(rows)
        self._count_lbl.setText(
            f"{count} lượt check-in" if count else "Chưa có lịch sử check-in."
        )
        self._table.setSortingEnabled(True)

    @Slot()
    def _on_member_changed(self) -> None:
        pass

    def selected_member_id(self) -> int | None:
        idx = self._combo.currentIndex()
        if idx < 0:
            return None
        return self._combo.itemData(idx)


# ══════════════════════════════════════════════════════════════════════════ #
#  ChartDialog  — popup riêng chứa Pie + Bar chart                           #
# ══════════════════════════════════════════════════════════════════════════ #

# ══════════════════════════════════════════════════════════════════════════ #
#  ChartDialog  — popup riêng chứa Pie + Bar chart                           #
# ══════════════════════════════════════════════════════════════════════════ #

class ChartDialog(QDialog):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("📈  Biểu Đồ Thống Kê Tháng Này")
        self.setMinimumSize(900, 480)
        self.resize(1000, 520)
        self.setModal(True)
        self._setup_ui()

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(12)

        month_str = date.today().strftime("%m/%Y")
        title_lbl = QLabel(f"Thống kê tháng {month_str}")
        title_lbl.setObjectName("sectionTitle")
        root.addWidget(title_lbl)

        charts_row = QHBoxLayout()
        charts_row.setSpacing(24)

        # ── SỬ DỤNG TRỰC TIẾP WIDGET ĐÃ CÓ TỪ CHART_WIDGET.PY ──
        self._pie_widget = PieChartWidget()
        charts_row.addWidget(self._pie_widget, stretch=1)

        vline = QFrame()
        vline.setFrameShape(QFrame.VLine)
        vline.setObjectName("sidebarDivider")
        charts_row.addWidget(vline)

        self._bar_widget = BarChartWidget()
        charts_row.addWidget(self._bar_widget, stretch=1)

        root.addLayout(charts_row)

        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_box.rejected.connect(self.reject)
        root.addWidget(btn_box, alignment=Qt.AlignRight)

    def load_data(self, checkin_status: dict, top_members: list[dict]) -> None:
        # Đẩy dữ liệu thẳng vào 2 widget để chúng tự vẽ UI hiện đại
        self._pie_widget.update_chart(checkin_status)
        self._bar_widget.update_chart(top_members)

    # (Đã xóa hàm _draw_pie và _draw_bar cũ ở đây vì không cần thiết nữa)

# ══════════════════════════════════════════════════════════════════════════ #
#  DashboardPage                                                              #
# ══════════════════════════════════════════════════════════════════════════ #

class DashboardPage(QWidget):

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._ctrl = DashboardController()
        self._cached_checkin_status: dict       = {"valid": 0, "expired": 0}
        self._cached_top_members:    list[dict] = []
        self._setup_ui()
        self.refresh_data()

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 12, 16, 12)
        root.setSpacing(12)

        root.addLayout(self._build_header())

        div = QFrame()
        div.setFrameShape(QFrame.HLine)
        div.setObjectName("sidebarDivider")
        root.addWidget(div)

        body_layout = QVBoxLayout()
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(12)

        # ── 1. Stat cards ──
        body_layout.addLayout(self._build_stat_cards())

        # ── 2. Today detail + Top members (chia đều 50/50) ──
        mid_row = QHBoxLayout()
        mid_row.setSpacing(16)

        self._today_table = TodayDetailTable()
        mid_row.addWidget(self._today_table, stretch=1)

        self._top_table = TopMembersTable()
        mid_row.addWidget(self._top_table, stretch=1)

        body_layout.addLayout(mid_row, stretch=1)

        # ── 3. Member history ──
        self._history_panel = MemberHistoryPanel()
        self._history_panel._combo.currentIndexChanged.connect(self._on_history_member_changed)
        body_layout.addWidget(self._history_panel, stretch=1)

        root.addLayout(body_layout, stretch=1)

        self._status_lbl = QLabel("")
        self._status_lbl.setObjectName("statusLabel")
        root.addWidget(self._status_lbl)

    def _build_header(self) -> QHBoxLayout:
        hdr = QHBoxLayout()
        hdr.setSpacing(14)

        icon = QLabel("📊")
        icon.setObjectName("pageIcon")

        title = QLabel("Dashboard")
        title.setObjectName("pageTitle")

        today_str = date.today().strftime("%d/%m/%Y")
        sub = QLabel(f"Tổng quan hoạt động phòng gym  •  {today_str}")
        sub.setObjectName("pageSubtitle")

        text_col = QVBoxLayout()
        text_col.setSpacing(2)
        text_col.addWidget(title)
        text_col.addWidget(sub)

        hdr.addWidget(icon)
        hdr.addLayout(text_col)
        hdr.addStretch()

        btn_export = QPushButton("📊  Xuất Báo Cáo")
        btn_export.setObjectName("btnNeutral")
        btn_export.setMinimumHeight(34)
        btn_export.setCursor(Qt.PointingHandCursor)
        btn_export.clicked.connect(self._export_excel_report)
        hdr.addWidget(btn_export)

        btn_chart = QPushButton("📈  Xem Biểu Đồ")
        btn_chart.setObjectName("btnNeutral")
        btn_chart.setMinimumHeight(34)
        btn_chart.setCursor(Qt.PointingHandCursor)
        btn_chart.clicked.connect(self._open_chart_dialog)
        hdr.addWidget(btn_chart)

        btn_refresh = QPushButton("🔄  Làm Mới")
        btn_refresh.setObjectName("btnNeutral")
        btn_refresh.setMinimumHeight(34)
        btn_refresh.setCursor(Qt.PointingHandCursor)
        btn_refresh.clicked.connect(self.refresh_data)
        hdr.addWidget(btn_refresh)

        return hdr

    def _build_stat_cards(self) -> QGridLayout:
        grid = QGridLayout()
        grid.setSpacing(12)

        self._card_today_total    = StatCard("🏃", "Tổng Check-in Hôm Nay",    "—", "#3b82f6")
        self._card_today_people   = StatCard("👤", "Số Người Đi Tập Hôm Nay",  "—", "#10b981")
        self._card_month_revenue  = StatCard("💰", "Doanh Thu Tháng Này",       "—", "#8b5cf6")
        self._card_year_revenue   = StatCard("📈", "Doanh Thu Cả Năm",          "—", "#f97316")
        self._card_month_people   = StatCard("👥", "Số Người Đi Tập Tháng Này","—", "#f59e0b")
        self._card_active         = StatCard("✅", "Hội Viên Đang Active",      "—", "#22c55e")

        grid.addWidget(self._card_today_total,   0, 0)
        grid.addWidget(self._card_today_people,  0, 1)
        grid.addWidget(self._card_month_revenue, 0, 2)
        grid.addWidget(self._card_year_revenue,  0, 3)
        grid.addWidget(self._card_month_people,  0, 4)
        grid.addWidget(self._card_active,        0, 5)

        for col in range(6):
            grid.setColumnStretch(col, 1)

        return grid

    def refresh_data(self) -> None:
        try:
            today_stats    = self._ctrl.get_today_stats()
            month_stats    = self._ctrl.get_month_stats()
            active_count   = self._ctrl.get_active_members()
            top_members    = self._ctrl.get_top_members(limit=5)
            all_members    = self._ctrl.get_all_members()
            checkin_status = self._ctrl.get_checkin_status_stats()

            self._cached_checkin_status = checkin_status
            self._cached_top_members    = top_members

            self._card_today_total.set_value(str(today_stats["total_checkins"]))
            self._card_today_people.set_value(str(today_stats["unique_members"]))

            month_rev = self._ctrl.get_month_revenue()
            year_rev  = self._ctrl.get_year_revenue()
            self._card_month_revenue.set_value(f"{month_rev:,.0f} ₫")
            self._card_year_revenue.set_value(f"{year_rev:,.0f} ₫")

            self._card_month_people.set_value(str(month_stats["unique_members"]))
            self._card_active.set_value(str(active_count))

            self._today_table.load(today_stats["detail"])
            self._top_table.load(top_members)

            current_id = self._history_panel.selected_member_id()
            self._history_panel.populate_members(all_members)

            if current_id is not None:
                combo = self._history_panel._combo
                for i in range(combo.count()):
                    if combo.itemData(i) == current_id:
                        combo.setCurrentIndex(i)
                        break
                self._load_member_history()
            elif all_members:
                self._history_panel._combo.setCurrentIndex(0)
                self._load_member_history()

            now_str = date.today().strftime("%d/%m/%Y")
            self._status_lbl.setText(f"Cập nhật lúc: {now_str}")

        except Exception as exc:
            self._status_lbl.setText(f"❌  Lỗi tải dữ liệu: {exc}")

    def _load_member_history(self) -> None:
        member_id = self._history_panel.selected_member_id()
        if member_id is None:
            self._history_panel.load_history([])
            return
        try:
            rows = self._ctrl.get_member_history(member_id)
            self._history_panel.load_history(rows)
        except Exception as exc:
            self._history_panel.load_history([])
            self._status_lbl.setText(f"❌  Lỗi tải lịch sử: {exc}")

    @Slot()
    def _open_chart_dialog(self) -> None:
        dlg = ChartDialog(self)
        dlg.load_data(self._cached_checkin_status, self._cached_top_members)
        dlg.exec()

    @Slot()
    def _on_history_member_changed(self) -> None:
        self._load_member_history()

    # ── Xuất báo cáo Excel ──────────────────────────────────────────────── #

    @Slot()
    def _export_excel_report(self) -> None:
        """Mở QFileDialog để chọn vị trí lưu, rồi xuất báo cáo Excel 2 sheet."""
        today_str = date.today().strftime("%Y-%m-%d")
        default_name = f"BaoCao_GymTK_{today_str}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Lưu Báo Cáo Excel",
            default_name,
            "Excel Files (*.xlsx);;All Files (*)",
        )

        if not file_path:
            return  # Người dùng huỷ

        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            self._status_lbl.setText("⏳  Đang xuất báo cáo...")
            self._write_excel_report(file_path)
            self._status_lbl.setText(f"✅  Đã xuất báo cáo: {file_path}")
            QMessageBox.information(
                self,
                "Xuất Báo Cáo Thành Công",
                f"Báo cáo đã được lưu tại:\n{file_path}",
            )
        except Exception as exc:
            self._status_lbl.setText(f"❌  Lỗi xuất báo cáo: {exc}")
            QMessageBox.critical(
                self,
                "Lỗi Xuất Báo Cáo",
                f"Không thể xuất báo cáo:\n{exc}",
            )

    def _write_excel_report(self, file_path: str) -> None:
        """Tạo file Excel với 2 sheet: Doanh Thu Tháng + Hội Viên Sắp Hết Hạn."""
        revenue_rows  = self._ctrl.get_month_revenue_detail()
        expiring_rows = self._ctrl.get_expiring_members(days_ahead=7)

        month_str = date.today().strftime("%m/%Y")

        # ── Sheet 1: Doanh Thu Tháng ──────────────────────────────────── #
        df_revenue = pd.DataFrame(
            revenue_rows,
            columns=["member_name", "plan_name", "price", "start_date"],
        ) if revenue_rows else pd.DataFrame(
            columns=["member_name", "plan_name", "price", "start_date"]
        )
        df_revenue.rename(columns={
            "member_name": "Họ Tên Hội Viên",
            "plan_name":   "Loại Gói",
            "price":       "Giá Tiền (VNĐ)",
            "start_date":  "Ngày Đăng Ký",
        }, inplace=True)

        # ── Sheet 2: Hội Viên Sắp Hết Hạn ────────────────────────────── #
        df_expiring = pd.DataFrame(
            expiring_rows,
            columns=["member_name", "phone", "end_date", "days_left"],
        ) if expiring_rows else pd.DataFrame(
            columns=["member_name", "phone", "end_date", "days_left"]
        )
        df_expiring.rename(columns={
            "member_name": "Họ Tên",
            "phone":       "Số Điện Thoại",
            "end_date":    "Ngày Hết Hạn",
            "days_left":   "Số Ngày Còn Lại",
        }, inplace=True)

        # ── Ghi ra file Excel với pandas ──────────────────────────────── #
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_revenue.to_excel(
                writer,
                sheet_name="Doanh Thu Tháng",
                index=False,
                startrow=2,       # Dành dòng 1–2 cho tiêu đề
            )
            df_expiring.to_excel(
                writer,
                sheet_name="Hội Viên Sắp Hết Hạn",
                index=False,
                startrow=2,
            )

        # ── Định dạng với openpyxl ────────────────────────────────────── #
        wb = load_workbook(file_path)
        self._format_sheet_revenue(wb["Doanh Thu Tháng"],  df_revenue,  month_str)
        self._format_sheet_expiring(wb["Hội Viên Sắp Hết Hạn"], df_expiring, month_str)
        wb.save(file_path)

    # ── Helpers định dạng Excel ─────────────────────────────────────────── #

    @staticmethod
    def _header_style() -> tuple:
        """Trả về (Font, PatternFill, Alignment) cho dòng tiêu đề cột."""
        font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        fill  = PatternFill("solid", start_color="1E3A5F")
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return font, fill, align

    @staticmethod
    def _title_style() -> tuple:
        """Trả về (Font, PatternFill, Alignment) cho dòng tiêu đề sheet."""
        font  = Font(name="Arial", bold=True, color="FFFFFF", size=13)
        fill  = PatternFill("solid", start_color="2563EB")
        align = Alignment(horizontal="center", vertical="center")
        return font, fill, align

    @staticmethod
    def _thin_border() -> Border:
        thin = Side(style="thin", color="D1D5DB")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _format_sheet_revenue(
        self,
        ws,
        df: pd.DataFrame,
        month_str: str,
    ) -> None:
        num_cols = len(df.columns)

        # ── Dòng tiêu đề sheet (row 1) ──
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws.cell(row=1, column=1, value=f"BÁO CÁO DOANH THU THÁNG {month_str}")
        t_font, t_fill, t_align = self._title_style()
        title_cell.font      = t_font
        title_cell.fill      = t_fill
        title_cell.alignment = t_align
        ws.row_dimensions[1].height = 28

        # ── Dòng header cột (row 3, vì startrow=2 → pandas viết header ở row 3) ──
        h_font, h_fill, h_align = self._header_style()
        border = self._thin_border()
        header_row = 3
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font      = h_font
            cell.fill      = h_fill
            cell.alignment = h_align
            cell.border    = border
        ws.row_dimensions[header_row].height = 22

        # ── Dữ liệu ──
        data_start = header_row + 1
        data_end   = data_start + len(df) - 1

        # Màu xen kẽ cho dòng dữ liệu
        fill_even = PatternFill("solid", start_color="EFF6FF")
        fill_odd  = PatternFill("solid", start_color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center")
        align_left   = Alignment(horizontal="left",   vertical="center")
        align_right  = Alignment(horizontal="right",  vertical="center")

        price_col_idx = 3  # Cột "Giá Tiền (VNĐ)"

        for row_idx in range(data_start, data_end + 1):
            row_fill = fill_even if (row_idx % 2 == 0) else fill_odd
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill   = row_fill
                cell.border = border
                if col_idx == price_col_idx:
                    cell.alignment  = align_right
                    cell.number_format = '#,##0'
                elif col_idx == num_cols:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
            ws.row_dimensions[row_idx].height = 18

        # ── Dòng tổng cộng ──
        if len(df) > 0:
            total_row = data_end + 1
            ws.merge_cells(
                start_row=total_row, start_column=1,
                end_row=total_row,   end_column=price_col_idx - 1,
            )
            total_label = ws.cell(row=total_row, column=1, value="TỔNG CỘNG")
            total_label.font      = Font(name="Arial", bold=True, size=11)
            total_label.alignment = Alignment(horizontal="right", vertical="center")
            total_label.border    = border

            total_val = ws.cell(
                row=total_row, column=price_col_idx,
                value=f"=SUM({get_column_letter(price_col_idx)}{data_start}:{get_column_letter(price_col_idx)}{data_end})",
            )
            total_val.font         = Font(name="Arial", bold=True, size=11, color="1D4ED8")
            total_val.alignment    = align_right
            total_val.number_format = '#,##0'
            total_val.border       = border
            ws.row_dimensions[total_row].height = 20

        # ── Độ rộng cột ──
        col_widths = [28, 22, 18, 16]
        for idx, width in enumerate(col_widths[:num_cols], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _format_sheet_expiring(
        self,
        ws,
        df: pd.DataFrame,
        month_str: str,
    ) -> None:
        num_cols = len(df.columns)

        # ── Tiêu đề sheet (row 1) ──
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws.cell(row=1, column=1, value="HỘI VIÊN SẮP HẾT HẠN (TRONG 7 NGÀY TỚI)")
        t_font, t_fill, t_align = self._title_style()
        title_cell.font      = t_font
        title_cell.fill      = t_fill
        title_cell.alignment = t_align
        ws.row_dimensions[1].height = 28

        # ── Header cột (row 3) ──
        h_font, h_fill, h_align = self._header_style()
        border = self._thin_border()
        header_row = 3
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font      = h_font
            cell.fill      = h_fill
            cell.alignment = h_align
            cell.border    = border
        ws.row_dimensions[header_row].height = 22

        # ── Dữ liệu ──
        data_start = header_row + 1
        data_end   = data_start + len(df) - 1

        fill_even    = PatternFill("solid", start_color="FFF7ED")
        fill_urgent  = PatternFill("solid", start_color="FEE2E2")   # ≤ 2 ngày: đỏ nhạt
        fill_warning = PatternFill("solid", start_color="FEF9C3")   # 3–5 ngày: vàng nhạt
        fill_odd     = PatternFill("solid", start_color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center")
        align_left   = Alignment(horizontal="left",   vertical="center")

        days_left_col = 4  # Cột "Số Ngày Còn Lại"

        for row_idx in range(data_start, data_end + 1):
            days_cell = ws.cell(row=row_idx, column=days_left_col)
            try:
                days_val = int(days_cell.value) if days_cell.value is not None else 99
            except (ValueError, TypeError):
                days_val = 99

            if days_val <= 2:
                row_fill = fill_urgent
            elif days_val <= 5:
                row_fill = fill_warning
            elif row_idx % 2 == 0:
                row_fill = fill_even
            else:
                row_fill = fill_odd

            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill   = row_fill
                cell.border = border
                cell.alignment = align_center if col_idx in (3, 4) else align_left
                if col_idx == days_left_col and days_val <= 2:
                    cell.font = Font(name="Arial", bold=True, color="DC2626", size=10)
            ws.row_dimensions[row_idx].height = 18

        # ── Độ rộng cột ──
        col_widths = [28, 18, 16, 18]
        for idx, width in enumerate(col_widths[:num_cols], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width